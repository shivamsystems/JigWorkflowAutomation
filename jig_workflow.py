
"""
===============================================================
  JIG WORKFLOW AUTOMATION v3 — FINAL
  ────────────────────────────────────
  • Auto-starts with Windows (no run.bat needed)
  • Monitors SolidWorks & Bambu Studio openings/closings
  • Auto-creates project folders
  • 5:15 PM daily handover popup with smart skip/remind logic
  • Logs everything to Excel
===============================================================
"""

import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
import os
import sys
import json
import threading
import time
import logging
import traceback
import argparse

# ── Core Imports ──────────────────────────────────────────────
import psutil

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("FATAL: openpyxl missing. Run: pip install openpyxl")
    sys.exit(1)

try:
    from pystray import Icon, Menu, MenuItem
    from PIL import Image, ImageDraw
    HAS_SYSTRAY = True
except ImportError:
    HAS_SYSTRAY = False

# ── CLI Arguments ─────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--startup", action="store_true", help="Auto-started by Windows")
cmd_args, _ = parser.parse_known_args()

# ── Paths & Logging ──────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_FILE = os.path.join(SCRIPT_DIR, "jig_workflow.log")
SETTINGS_FILE = os.path.join(SCRIPT_DIR, "settings.json")

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, mode="a", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("JigWorkflow")

# ── Defaults ─────────────────────────────────────────────────
DEFAULT_SETTINGS = {
    "excel_file": os.path.join(SCRIPT_DIR, "Jig_Workflow_Log.xlsx"),
    "projects_base_folder": os.path.join(os.path.expanduser("~"), "JigProjects"),
    "folder_subfolders": ["CAD", "STL", "Prints", "References", "Notes"],
    "sw_process_names": ["sldworks.exe", "SLDWORKS.EXE", "SolidWorks.exe"],
    "bambu_process_names": ["bambustudio.exe", "BambuStudio.exe", "orcaslicer.exe"],
    "check_interval_sec": 3,
    "auto_create_folders": True,
    "handover_hour": 17,          # 5 PM
    "handover_minute": 15,        # 5:15 PM
    "handover_deadline_hour": 18, # 6 PM
}


# ============================================================
#  SYSTEM 1: SETTINGS MANAGER
# ============================================================
class SettingsManager:
    def __init__(self, filepath):
        self.filepath = filepath
        self.settings = dict(DEFAULT_SETTINGS)
        self.load()

    def load(self):
        if os.path.exists(self.filepath):
            try:
                with open(self.filepath, "r", encoding="utf-8") as f:
                    saved = json.load(f)
                self.settings.update(saved)
            except Exception as e:
                log.error(f"Settings load error: {e}")

    def save(self):
        try:
            with open(self.filepath, "w", encoding="utf-8") as f:
                json.dump(self.settings, f, indent=4)
        except Exception as e:
            log.error(f"Settings save error: {e}")

    def get(self, key, default=None):
        return self.settings.get(key, default)

    def set(self, key, value):
        self.settings[key] = value
        self.save()


# ============================================================
#  SYSTEM 2: EXCEL LOGGER (Thread-Safe with Retry)
# ============================================================
class ExcelLogger:
    SW_SHEET = "SolidWorks Projects"
    PRINT_SHEET = "Print & Filament Log"
    HANDOVER_SHEET = "Daily Handovers"

    def __init__(self, filepath):
        self.filepath = filepath
        self._lock = threading.Lock()
        self._init_workbook()

    def _init_workbook(self):
        if os.path.exists(self.filepath):
            return
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        self._create_sw_sheet(wb)
        self._create_print_sheet(wb)
        self._create_handover_sheet(wb)
        wb.save(self.filepath)

    def _style_headers(self, ws, count):
        fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _create_sw_sheet(self, wb):
        ws = wb.create_sheet(self.SW_SHEET)
        headers = ["Date", "Start Time", "End Time", "Project Name", "Jig / Assembly Name", 
                   "Description", "Priority", "Status", "Work Completed", "Duration (min)", "Project Folder"]
        ws.append(headers)
        self._style_headers(ws, len(headers))
        widths = [12, 10, 10, 24, 24, 36, 12, 14, 36, 14, 36]
        for i, w in enumerate(widths, 1): 
            ws.column_dimensions[get_column_letter(i)].width = w

    def _create_print_sheet(self, wb):
        ws = wb.create_sheet(self.PRINT_SHEET)
        headers = ["Date", "Time", "Filament Incoming (g)", "Model Printed", "Model Used For", 
                   "Material Consumed (g)", "Filament Type", "Color", "Print Result", "Notes"]
        ws.append(headers)
        self._style_headers(ws, len(headers))
        widths = [12, 10, 18, 28, 22, 20, 14, 12, 14, 36]
        for i, w in enumerate(widths, 1): 
            ws.column_dimensions[get_column_letter(i)].width = w

    def _create_handover_sheet(self, wb):
        ws = wb.create_sheet(self.HANDOVER_SHEET)
        headers = ["Date", "Time", "Item Received", "Received From", "Item Returned / Handed Over", 
                   "Handed Over To", "Jig Involved", "Status", "Notes"]
        ws.append(headers)
        self._style_headers(ws, len(headers))
        widths = [12, 10, 28, 20, 28, 20, 20, 14, 36]
        for i, w in enumerate(widths, 1): 
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_with_retry(self, operation, retries=3):
        for attempt in range(retries):
            try:
                return operation()
            except PermissionError:
                log.warning(f"Excel locked, retry {attempt+1}...")
                time.sleep(2)
            except Exception as e:
                log.error(f"Excel write error: {e}")
                return None
        return None

    def get_recent_projects(self, limit=15):
        try:
            wb = openpyxl.load_workbook(self.filepath, read_only=True)
            ws = wb[self.SW_SHEET]
            names, seen = [], set()
            for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
                name = row[3]
                if name and str(name) not in seen:
                    seen.add(str(name))
                    names.append(str(name))
            wb.close()
            return names[:limit]
        except Exception:
            return []

    def log_sw_start(self, data):
        def _w():
            wb = openpyxl.load_workbook(self.filepath)
            ws = wb[self.SW_SHEET]
            ws.append([
                datetime.now().strftime("%Y-%m-%d"), 
                data.get("start_time", ""), 
                "", 
                data.get("project_name", ""), 
                data.get("jig_name", ""), 
                data.get("description", ""), 
                data.get("priority", ""), 
                "In Progress", 
                "", 
                "", 
                data.get("project_folder", "")
            ])
            r = ws.max_row
            wb.save(self.filepath)
            wb.close()
            return r
        with self._lock:
            return self._write_with_retry(_w)

    def update_sw_end(self, row_num, data):
        if not row_num:
            return False
            
        def _w():
            wb = openpyxl.load_workbook(self.filepath)
            ws = wb[self.SW_SHEET]
            if data.get("end_time"): 
                ws.cell(row=row_num, column=3).value = data["end_time"]
            if data.get("status"): 
                ws.cell(row=row_num, column=8).value = data["status"]
            if data.get("work_completed"): 
                ws.cell(row=row_num, column=9).value = data["work_completed"]
            if data.get("duration_min") is not None: 
                ws.cell(row=row_num, column=10).value = data["duration_min"]
            wb.save(self.filepath)
            wb.close()
            return True
        with self._lock:
            return self._write_with_retry(_w)

    def log_print(self, data):
        def _w():
            wb = openpyxl.load_workbook(self.filepath)
            wb[self.PRINT_SHEET].append([
                datetime.now().strftime("%Y-%m-%d"), 
                datetime.now().strftime("%H:%M:%S"), 
                data.get("filament_incoming", ""), 
                data.get("model_printed", ""), 
                data.get("model_used", ""), 
                data.get("material_consumed", ""), 
                data.get("filament_type", ""), 
                data.get("color", ""), 
                data.get("print_result", ""), 
                data.get("notes", "")
            ])
            wb.save(self.filepath)
            wb.close()
            return True
        with self._lock:
            return self._write_with_retry(_w)

    def log_handover(self, data):
        def _w():
            wb = openpyxl.load_workbook(self.filepath)
            wb[self.HANDOVER_SHEET].append([
                datetime.now().strftime("%Y-%m-%d"), 
                datetime.now().strftime("%H:%M:%S"), 
                data.get("item_received", ""), 
                data.get("received_from", ""), 
                data.get("item_returned", ""), 
                data.get("handed_over_to", ""), 
                data.get("jig_involved", ""), 
                data.get("status", ""), 
                data.get("notes", "")
            ])
            wb.save(self.filepath)
            wb.close()
            return True
        with self._lock:
            return self._write_with_retry(_w)


# ============================================================
#  SYSTEM 3: FOLDER CREATOR
# ============================================================
class FolderCreator:
    def __init__(self, settings):
        self.base_folder = settings.get("projects_base_folder")
        self.subfolders = settings.get("folder_subfolders", [])

    def create_project_structure(self, project_name, jig_name=""):
        if not project_name:
            return ""
            
        parts = [self.base_folder, self._safe_name(project_name)]
        if jig_name:
            parts.append(self._safe_name(jig_name))
            
        project_path = os.path.join(*parts)
        try:
            os.makedirs(project_path, exist_ok=True)
            for sub in self.subfolders:
                os.makedirs(os.path.join(project_path, sub), exist_ok=True)
            return project_path
        except Exception as e:
            log.error(f"Folder error: {e}")
            return ""

    @staticmethod
    def _safe_name(name):
        for c in '<>:"/\\|?*':
            name = name.replace(c, "_")
        return name.strip()


# ============================================================
#  SYSTEM 4: USER INTERFACES (POPUPS)
# ============================================================
class SWStartPopup:
    def __init__(self, parent, recent_projects=None, settings=None):
        self.data = None
        self.win = tk.Toplevel(parent)
        self.win.title("New Jig Project")
        self.win.geometry("540x480")
        self.win.resizable(False, False)
        self.win.attributes("-topmost", True)
        self.win.grab_set()
        self.win.configure(bg="#F0F4FA")

        # Header
        tk.Label(self.win, text="  Start New Jig Session", font=("Segoe UI", 15, "bold"), bg="#2F5496", fg="white").pack(fill="x", ipady=10)
        
        # Body
        body = tk.Frame(self.win, bg="#F0F4FA", padx=25, pady=12)
        body.pack(fill="both", expand=True)

        # Project Name
        tk.Label(body, text="Project Name *", font=("Segoe UI", 10, "bold"), bg="#F0F4FA").grid(row=0, column=0, sticky="w", pady=6)
        self.project_name = ttk.Combobox(body, values=recent_projects or [], width=36)
        self.project_name.grid(row=0, column=1, pady=6, padx=(12, 0))

        # Jig Name
        tk.Label(body, text="Jig / Assembly Name", font=("Segoe UI", 10), bg="#F0F4FA").grid(row=1, column=0, sticky="w", pady=6)
        self.jig_name = ttk.Entry(body, width=39)
        self.jig_name.grid(row=1, column=1, pady=6, padx=(12, 0))

        # Description
        tk.Label(body, text="Description", font=("Segoe UI", 10), bg="#F0F4FA").grid(row=2, column=0, sticky="nw", pady=6)
        self.description = tk.Text(body, width=39, height=3, font=("Segoe UI", 10))
        self.description.grid(row=2, column=1, pady=6, padx=(12, 0))

        # Priority
        tk.Label(body, text="Priority", font=("Segoe UI", 10), bg="#F0F4FA").grid(row=3, column=0, sticky="w", pady=6)
        self.priority = ttk.Combobox(body, values=["Low", "Medium", "High", "Urgent"], state="readonly", width=36)
        self.priority.set("Medium")
        self.priority.grid(row=3, column=1, pady=6, padx=(12, 0))

        # Auto Folder Checkbox
        self.auto_folder = tk.BooleanVar(value=settings.get("auto_create_folders", True) if settings else True)
        tk.Checkbutton(body, text="Auto-create project folder structure", variable=self.auto_folder, font=("Segoe UI", 10), bg="#F0F4FA").grid(row=4, column=0, columnspan=2, sticky="w", pady=8)

        # Buttons
        btn_frame = tk.Frame(body, bg="#F0F4FA")
        btn_frame.grid(row=5, column=0, columnspan=2, pady=(18, 0))
        tk.Button(btn_frame, text="Start Session", font=("Segoe UI", 11, "bold"), bg="#2F5496", fg="white", padx=22, pady=7, command=self._submit).pack(side="right", padx=(8, 0))
        tk.Button(btn_frame, text="Skip", font=("Segoe UI", 10), padx=14, pady=7, command=self._skip).pack(side="right")
        
        self.win.protocol("WM_DELETE_WINDOW", self._skip)

    def _submit(self):
        if not self.project_name.get().strip():
            messagebox.showwarning("Required", "Please enter a Project Name.", parent=self.win)
            return
        self.data = {
            "project_name": self.project_name.get().strip(), 
            "jig_name": self.jig_name.get().strip(), 
            "description": self.description.get("1.0", "end").strip(), 
            "priority": self.priority.get(), 
            "start_time": datetime.now().strftime("%H:%M:%S"), 
            "auto_create_folder": self.auto_folder.get()
        }
        self.win.destroy()

    def _skip(self):
        self.data = None
        self.win.destroy()


class SWEndPopup:
    def __init__(self, parent, project_data):
        self.data = None
        self.start_time_obj = project_data.get("start_time_obj", datetime.now())
        
        self.win = tk.Toplevel(parent)
        self.win.title("Session Complete")
        self.win.geometry("520x380")
        self.win.resizable(False, False)
        self.win.attributes("-topmost", True)
        self.win.grab_set()
        self.win.configure(bg="#F0F4FA")

        # Header
        tk.Label(self.win, text=f"  Session: {project_data.get('project_name', 'Unknown')}", font=("Segoe UI", 14, "bold"), bg="#548235", fg="white").pack(fill="x", ipady=10)
        
        # Body
        body = tk.Frame(self.win, bg="#F0F4FA", padx=25, pady=15)
        body.pack(fill="both", expand=True)

        # Work Completed
        tk.Label(body, text="Work Completed", font=("Segoe UI", 10, "bold"), bg="#F0F4FA").grid(row=0, column=0, sticky="nw", pady=6)
        self.work_completed = tk.Text(body, width=39, height=5, font=("Segoe UI", 10))
        self.work_completed.grid(row=0, column=1, pady=6, padx=(12, 0))

        # Status
        tk.Label(body, text="Status", font=("Segoe UI", 10, "bold"), bg="#F0F4FA").grid(row=1, column=0, sticky="w", pady=6)
        self.status = ttk.Combobox(body, values=["In Progress", "Complete", "On Hold", "Cancelled"], state="readonly", width=36)
        self.status.set("In Progress")
        self.status.grid(row=1, column=1, pady=6, padx=(12, 0))

        # Buttons
        btn_frame = tk.Frame(body, bg="#F0F4FA")
        btn_frame.grid(row=2, column=0, columnspan=2, pady=(20, 0))
        tk.Button(btn_frame, text="Save & Close", font=("Segoe UI", 11, "bold"), bg="#548235", fg="white", padx=22, pady=7, command=self._submit).pack(side="right", padx=(8, 0))
        tk.Button(btn_frame, text="Skip", font=("Segoe UI", 10), padx=14, pady=7, command=self._skip).pack(side="right")
        
        self.win.protocol("WM_DELETE_WINDOW", self._skip)

    def _submit(self):
        et = datetime.now()
        d = et - self.start_time_obj
        self.data = {
            "end_time": et.strftime("%H:%M:%S"), 
            "status": self.status.get(), 
            "work_completed": self.work_completed.get("1.0", "end").strip(), 
            "duration_min": round(d.total_seconds() / 60, 1)
        }
        self.win.destroy()

    def _skip(self):
        self.data = None
        self.win.destroy()


class PrintLogPopup:
    def __init__(self, parent):
        self.data = None
        self.win = tk.Toplevel(parent)
        self.win.title("Print & Filament Log")
        self.win.geometry("540x600")
        self.win.resizable(False, False)
        self.win.attributes("-topmost", True)
        self.win.grab_set()
        self.win.configure(bg="#F0F4FA")

        # Header
        tk.Label(self.win, text="  Log Print / Filament", font=("Segoe UI", 14, "bold"), bg="#BF8F00", fg="white").pack(fill="x", ipady=10)

        # Scrollable Canvas
        canvas = tk.Canvas(self.win, bg="#F0F4FA", highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.win, orient="vertical", command=canvas.yview)
        body = tk.Frame(canvas, bg="#F0F4FA", padx=25, pady=15)
        body.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        r = 0
        def field(lbl, row, w_type="entry", vals=None, ro=False, bold=False):
            f = ("Segoe UI", 10, "bold") if bold else ("Segoe UI", 10)
            tk.Label(body, text=lbl, font=f, bg="#F0F4FA").grid(row=row, column=0, sticky="w", pady=4)
            if w_type == "entry": w = ttk.Entry(body, width=37)
            elif w_type == "combo": w = ttk.Combobox(body, values=vals or [], state="readonly" if ro else "normal", width=34)
            elif w_type == "text": w = tk.Text(body, width=37, height=3, font=("Segoe UI", 10))
            else: w = ttk.Entry(body, width=37)
            w.grid(row=row, column=1, pady=4, padx=(12, 0))
            return w

        self.model_printed = field("Model Printed *", r, bold=True); r += 1
        self.model_used = field("Model Used For", r); r += 1
        self.filament_type = field("Filament Type", r, w_type="combo", vals=["PLA", "PLA+", "PETG", "ABS", "ASA", "TPU", "Nylon", "PC", "Other"], ro=True)
        self.filament_type.set("PLA"); r += 1
        self.color = field("Color", r); r += 1
        self.filament_incoming = field("Filament Incoming (g)", r); r += 1
        self.material_consumed = field("Material Consumed (g)", r); r += 1
        self.print_result = field("Print Result", r, w_type="combo", vals=["Success", "Partial", "Failed"], ro=True)
        self.print_result.set("Success"); r += 1
        self.notes = field("Notes", r, w_type="text"); r += 1

        # Buttons
        btn_frame = tk.Frame(body, bg="#F0F4FA")
        btn_frame.grid(row=r, column=0, columnspan=2, pady=(18, 5))
        tk.Button(btn_frame, text="Save Log", font=("Segoe UI", 11, "bold"), bg="#BF8F00", fg="white", padx=22, pady=7, command=self._submit).pack(side="right", padx=(8, 0))
        tk.Button(btn_frame, text="Cancel", font=("Segoe UI", 10), padx=14, pady=7, command=self._cancel).pack(side="right")
        self.win.protocol("WM_DELETE_WINDOW", self._cancel)

    def _submit(self):
        if not self.model_printed.get().strip():
            messagebox.showwarning("Required", "Please enter Model Printed.", parent=self.win)
            return
        self.data = {
            "model_printed": self.model_printed.get().strip(), 
            "model_used": self.model_used.get().strip(), 
            "filament_type": self.filament_type.get(), 
            "color": self.color.get().strip(), 
            "filament_incoming": self.filament_incoming.get().strip(), 
            "material_consumed": self.material_consumed.get().strip(), 
            "print_result": self.print_result.get(), 
            "notes": self.notes.get("1.0", "end").strip()
        }
        self.win.destroy()

    def _cancel(self):
        self.data = None
        self.win.destroy()


class HandoverPopup:
    def __init__(self, parent):
        self.data = None
        self.win = tk.Toplevel(parent)
        self.win.title("Daily Handover Log")
        self.win.geometry("560x550")
        self.win.resizable(False, False)
        self.win.attributes("-topmost", True)
        self.win.grab_set()
        self.win.configure(bg="#F0F4FA")

        # Header
        tk.Label(self.win, text="  5:15 PM — Daily Handover", font=("Segoe UI", 14, "bold"), bg="#C00000", fg="white").pack(fill="x", ipady=10)
        
        # Body
        body = tk.Frame(self.win, bg="#F0F4FA", padx=25, pady=15)
        body.pack(fill="both", expand=True)

        r = 0
        def field(lbl, row, w_type="entry", vals=None, ro=False, bold=False):
            f = ("Segoe UI", 10, "bold") if bold else ("Segoe UI", 10)
            tk.Label(body, text=lbl, font=f, bg="#F0F4FA").grid(row=row, column=0, sticky="w", pady=5)
            if w_type == "entry": w = ttk.Entry(body, width=38)
            elif w_type == "combo": w = ttk.Combobox(body, values=vals or [], state="readonly" if ro else "normal", width=35)
            elif w_type == "text": w = tk.Text(body, width=38, height=2, font=("Segoe UI", 10))
            else: w = ttk.Entry(body, width=38)
            w.grid(row=row, column=1, pady=5, padx=(10, 0))
            return w

        # Received Section
        tk.Label(body, text="── RECEIVED ──", font=("Segoe UI", 10, "bold"), fg="#2F5496", bg="#F0F4FA").grid(row=r, column=0, columnspan=2, pady=(5, 2)); r += 1
        self.item_received = field("Item Received *", r, bold=True); r += 1
        self.received_from = field("Received From", r); r += 1

        # Returned Section
        tk.Label(body, text="── RETURNED / HANDED OVER ──", font=("Segoe UI", 10, "bold"), fg="#548235", bg="#F0F4FA").grid(row=r, column=0, columnspan=2, pady=(10, 2)); r += 1
        self.item_returned = field("Item Returned", r); r += 1
        self.handed_over_to = field("Handed Over To", r); r += 1

        # Details Section
        tk.Label(body, text="── DETAILS ──", font=("Segoe UI", 10, "bold"), fg="#BF8F00", bg="#F0F4FA").grid(row=r, column=0, columnspan=2, pady=(10, 2)); r += 1
        self.jig_involved = field("Jig Involved", r); r += 1
        self.status = field("Status", r, w_type="combo", vals=["Complete", "Partial", "Pending", "Issue Found"], ro=True)
        self.status.set("Complete"); r += 1
        self.notes = field("Notes", r, w_type="text"); r += 1

        # Buttons
        btn_frame = tk.Frame(body, bg="#F0F4FA")
        btn_frame.grid(row=r, column=0, columnspan=2, pady=(15, 5))
        tk.Button(btn_frame, text="Save Handover", font=("Segoe UI", 11, "bold"), bg="#C00000", fg="white", padx=22, pady=7, command=self._submit).pack(side="right", padx=(8, 0))
        tk.Button(btn_frame, text="Skip", font=("Segoe UI", 10), padx=14, pady=7, command=self._skip).pack(side="right")
        
        self.win.protocol("WM_DELETE_WINDOW", self._skip)

    def _submit(self):
        if not self.item_received.get().strip():
            messagebox.showwarning("Required", "What did you receive today?", parent=self.win)
            return
        self.data = {
            "item_received": self.item_received.get().strip(), 
            "received_from": self.received_from.get().strip(), 
            "item_returned": self.item_returned.get().strip(), 
            "handed_over_to": self.handed_over_to.get().strip(), 
            "jig_involved": self.jig_involved.get().strip(), 
            "status": self.status.get(), 
            "notes": self.notes.get("1.0", "end").strip()
        }
        self.win.destroy()

    def _skip(self):
        self.data = None
        self.win.destroy()


# ============================================================
#  SYSTEM 5: MAIN APPLICATION BRAIN
# ============================================================
class App:
    def __init__(self):
        self.settings_mgr = SettingsManager(SETTINGS_FILE)
        self.settings = self.settings_mgr.settings
        self.logger = ExcelLogger(self.settings.get("excel_file"))
        self.folder_creator = FolderCreator(self.settings)

        # Process Tracking
        self.sw_pids = set()
        self.bambu_pids = set()
        self.active_sw_sessions = {}
        self.is_startup = getattr(cmd_args, "startup", False)
        
        # Handover tracking state machine
        self.handover_last_date = ""
        self.handover_status = "Waiting"  # "Waiting", "Pending", "Done"
        self.handover_remind_at = None
        self.handover_remind_minutes = 15 # Configurable reminder time

        # Main Window
        self.root = tk.Tk()
        self.root.title("Jig Workflow Automation v3")
        self.root.geometry("420x540")
        self.root.configure(bg="#F0F4FA")
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        self._build_dashboard()

        # If auto-started by Windows, hide to tray
        if self.is_startup and HAS_SYSTRAY:
            self.root.after(500, self._minimize_to_tray)

        # Initial scan
        self.sw_pids, self.bambu_pids = self._scan_processes()
        self._update_process_display()

        # Start background loops
        self._process_check_loop()
        self._time_check_loop()

        # System Tray
        self.tray_icon = None
        if HAS_SYSTRAY:
            self._setup_systray()

        self._add_log("Tool started. Monitoring active.")
        log.info("Jig Workflow Automation v3 started")
        
        # Next day startup check
        self.root.after(3000, self._check_missed_day)
        
        self.root.mainloop()

    # ── Dashboard UI ──────────────────────────────────────────
    def _build_dashboard(self):
        # Title
        tk.Label(self.root, text="  Jig Workflow Automation", font=("Segoe UI", 13, "bold"), bg="#2F5496", fg="white").pack(fill="x", ipady=8)
        
        # Status Frame
        status_frame = tk.LabelFrame(self.root, text="  Status  ", font=("Segoe UI", 10, "bold"), bg="#F0F4FA", padx=12, pady=8)
        status_frame.pack(fill="x", padx=12, pady=(10, 4))
        
        self.sw_status_label = tk.Label(status_frame, text="SolidWorks: Scanning...", font=("Segoe UI", 10), bg="#F0F4FA", anchor="w")
        self.sw_status_label.pack(fill="x", pady=2)
        
        self.bambu_status_label = tk.Label(status_frame, text="Bambu Studio: Scanning...", font=("Segoe UI", 10), bg="#F0F4FA", anchor="w")
        self.bambu_status_label.pack(fill="x", pady=2)
        
        self.session_label = tk.Label(status_frame, text="No active session", font=("Segoe UI", 10, "italic"), fg="#888", bg="#F0F4FA", anchor="w")
        self.session_label.pack(fill="x", pady=2)
        
        self.handover_label = tk.Label(status_frame, text="Handover: Waiting for 5:15 PM", font=("Segoe UI", 10, "italic"), fg="#888", bg="#F0F4FA", anchor="w")
        self.handover_label.pack(fill="x", pady=2)

        # Action Buttons Frame
        action_frame = tk.LabelFrame(self.root, text="  Quick Actions  ", font=("Segoe UI", 10, "bold"), bg="#F0F4FA", padx=12, pady=8)
        action_frame.pack(fill="x", padx=12, pady=4)
        
        btn_style = {"font": ("Segoe UI", 10), "padx": 10, "pady": 5, "width": 24, "anchor": "w"}
        tk.Button(action_frame, text="  Log New Project", bg="#2F5496", fg="white", command=self._manual_sw_log, **btn_style).pack(fill="x", pady=2)
        tk.Button(action_frame, text="  Log Print / Filament", bg="#BF8F00", fg="white", command=self._manual_print_log, **btn_style).pack(fill="x", pady=2)
        tk.Button(action_frame, text="  Log Handover Now", bg="#C00000", fg="white", command=self._manual_handover_log, **btn_style).pack(fill="x", pady=2)
        tk.Button(action_frame, text="  Open Excel Log", command=self._open_excel, **btn_style).pack(fill="x", pady=2)
        tk.Button(action_frame, text="  Open Project Folder", command=self._open_project_folder, **btn_style).pack(fill="x", pady=2)

        # Activity Log Frame
        log_frame = tk.LabelFrame(self.root, text="  Activity Log  ", font=("Segoe UI", 10, "bold"), bg="#F0F4FA", padx=8, pady=4)
        log_frame.pack(fill="both", expand=True, padx=12, pady=(4, 10))
        
        self.activity_log = tk.Text(log_frame, height=6, font=("Consolas", 9), bg="#FFFFFF", state="disabled", wrap="word")
        self.activity_log.pack(fill="both", expand=True)

    def _add_log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.activity_log.config(state="normal")
        self.activity_log.insert("1.0", f"[{ts}] {msg}\n")
        self.activity_log.config(state="disabled")

    def _update_process_display(self):
        self.sw_status_label.config(text=f"SolidWorks: {'RUNNING' if self.sw_pids else 'Not running'}", fg="#2E7D32" if self.sw_pids else "#999")
        self.bambu_status_label.config(text=f"Bambu Studio: {'RUNNING' if self.bambu_pids else 'Not running'}", fg="#2E7D32" if self.bambu_pids else "#999")
        
        if self.active_sw_sessions:
            s = list(self.active_sw_sessions.values())[0]
            self.session_label.config(text=f"Active: {s['data'].get('project_name', '')} (since {s['data'].get('start_time', '')})", fg="#2F5496", font=("Segoe UI", 10, "bold"))
        else:
            self.session_label.config(text="No active session", fg="#888", font=("Segoe UI", 10, "italic"))

    # ── Process Scanning (The Watcher) ────────────────────────
    def _scan_processes(self):
        sw, bambu = set(), set()
        sw_kw = [n.lower().replace(".exe", "") for n in self.settings.get("sw_process_names", [])]
        bambu_kw = [n.lower().replace(".exe", "") for n in self.settings.get("bambu_process_names", [])]
        
        for proc in psutil.process_iter(['pid', 'name']):
            try:
                nm = proc.info['name']
                pid = proc.info['pid']
                if not nm: continue
                nl = nm.lower()
                if any(k in nl for k in sw_kw): sw.add(pid)
                elif any(k in nl for k in bambu_kw): bambu.add(pid)
            except:
                pass
        return sw, bambu

    def _process_check_loop(self):
        try:
            cur_sw, cur_bambu = self._scan_processes()
            
            # SolidWorks Events
            for p in cur_sw - self.sw_pids:
                log.info(f"SW started PID {p}")
                self._add_log("SolidWorks detected")
                self._handle_sw_start(p)
            for p in self.sw_pids - cur_sw:
                log.info(f"SW closed PID {p}")
                self._add_log("SolidWorks closed")
                self._handle_sw_end(p)
                
            # Bambu Studio Events
            for p in self.bambu_pids - cur_bambu:
                log.info(f"Bambu closed PID {p}")
                self._add_log("Bambu Studio closed")
                self._handle_print_log()
                
            self.sw_pids = cur_sw
            self.bambu_pids = cur_bambu
            self._update_process_display()
        except Exception as e:
            log.error(f"Check loop error: {e}")
            
        # Check again in 3 seconds
        self.root.after(self.settings.get("check_interval_sec", 3) * 1000, self._process_check_loop)

    # ── Time Check Loop (The Clock State Machine) ─────────────
    def _time_check_loop(self):
        now = datetime.now()
        today_str = now.strftime("%Y-%m-%d")
        h_target = self.settings.get("handover_hour", 17)
        m_target = self.settings.get("handover_minute", 15)
        h_deadline = self.settings.get("handover_deadline_hour", 18)

        # STATE 1: Before 5:15 PM
        if now.hour < h_target or (now.hour == h_target and now.minute < m_target):
            self.handover_status = "Waiting"
            self.handover_label.config(text=f"Handover: Waiting for {h_target}:{m_target:02d}", fg="#888")

        # STATE 2: Between 5:15 PM and 6:00 PM (Action Time)
        elif (now.hour == h_target and now.minute >= m_target) or (now.hour > h_target and now.hour < h_deadline):
            if self.handover_status == "Done":
                pass # Do nothing, already filled
            elif self.handover_status == "Pending":
                if self.handover_remind_at and now >= self.handover_remind_at:
                    self._add_log("Handover reminder triggered!")
                    self._handle_handover_log() # Ask again
            elif self.handover_status == "Waiting":
                log.info("5:15 PM reached - triggering handover popup")
                self._add_log(f"{h_target}:{m_target:02d}: Time for daily handover log!")
                self.handover_status = "Pending"
                self.handover_last_date = today_str
                self._handle_handover_log()

        # STATE 3: Past 6:00 PM (Deadline)
        elif now.hour >= h_deadline:
            if self.handover_status != "Done":
                self._add_log("6:00 PM Deadline: Handover not filled today.")
                self.handover_label.config(text="Handover: Deadline Passed ⚠️", fg="red")
                self.handover_status = "Done" # Stop asking until tomorrow
            
        # Check again in 30 seconds
        self.root.after(30000, self._time_check_loop) 

    # ── Event Handlers ────────────────────────────────────────
    def _handle_sw_start(self, pid):
        popup = SWStartPopup(self.root, self.logger.get_recent_projects(), self.settings)
        self.root.wait_window(popup.win)
        
        if popup.data:
            pf = ""
            if popup.data.get("auto_create_folder"):
                pf = self.folder_creator.create_project_structure(popup.data["project_name"], popup.data.get("jig_name", ""))
                if pf: self._add_log(f"Folder: {pf}")
            
            popup.data["start_time_obj"] = datetime.now()
            popup.data["project_folder"] = pf
            rn = self.logger.log_sw_start(popup.data)
            self.active_sw_sessions[pid] = {"data": popup.data, "row_num": rn, "start_time_obj": popup.data["start_time_obj"]}
            self._add_log(f"Session: {popup.data['project_name']}")

    def _handle_sw_end(self, pid):
        session = self.active_sw_sessions.pop(pid, None)
        if not session: return
        
        popup = SWEndPopup(self.root, {"project_name": session["data"].get("project_name", ""), "start_time_obj": session["start_time_obj"]})
        self.root.wait_window(popup.win)
        
        if popup.data:
            self.logger.update_sw_end(session["row_num"], popup.data)
            self._add_log(f"Closed: {session['data'].get('project_name', '')} ({popup.data.get('duration_min', '?')}m)")
        else:
            et = datetime.now()
            d = et - session["start_time_obj"]
            self.logger.update_sw_end(session["row_num"], {
                "end_time": et.strftime("%H:%M:%S"), 
                "status": "In Progress", 
                "work_completed": "(Auto-closed)", 
                "duration_min": round(d.total_seconds() / 60, 1)
            })

    def _handle_print_log(self):
        popup = PrintLogPopup(self.root)
        self.root.wait_window(popup.win)
        if popup.data:
            self.logger.log_print(popup.data)
            self._add_log(f"Print: {popup.data.get('model_printed', '')}")

    def _handle_handover_log(self):
        remind_mins = self.handover_remind_minutes
        popup = HandoverPopup(self.root)
        self.root.wait_window(popup.win)
        
        if popup.data:
            self.logger.log_handover(popup.data)
            self._add_log("Handover logged successfully")
            self.handover_status = "Done" # Mark as Done, loop won't ask again today
            self.handover_label.config(text="Handover: Logged today ✓", fg="#2E7D32")
        else:
            self._add_log(f"Handover skipped. Reminding in {remind_mins}m.")
            self.handover_status = "Pending" # Mark as Pending
            self.handover_remind_at = datetime.now() + timedelta(minutes=remind_mins)
            self.handover_label.config(text=f"Handover: Skipped (Next in {remind_mins}m)", fg="#BF8F00")

    def _check_missed_day(self):
        """On startup, check if yesterday's handover was missed."""
        if self.handover_status != "Done" and self.handover_last_date != datetime.now().strftime("%Y-%m-%d"):
            answer = messagebox.askyesno(
                "Missed Handover", 
                "It looks like you didn't fill out yesterday's 5:15 PM Handover log.\n\nDo you want to fill it out now?", 
                parent=self.root
            )
            if answer:
                self._handle_handover_log()

    # ── Manual Buttons ────────────────────────────────────────
    def _manual_sw_log(self):
        self._handle_sw_start(0)

    def _manual_print_log(self):
        self._handle_print_log()

    def _manual_handover_log(self):
        self._handle_handover_log()

    def _open_excel(self):
        f = self.settings.get("excel_file")
        if os.path.exists(f): os.startfile(f)

    def _open_project_folder(self):
        b = self.settings.get("projects_base_folder")
        if b and os.path.exists(b): os.startfile(b)

    # ── System Tray ───────────────────────────────────────────
    def _minimize_to_tray(self):
        self.root.withdraw()

    def _restore_from_tray(self):
        self.root.deiconify()
        self.root.lift()
        self.root.focus_force()

    def _on_close(self):
        if self.tray_icon:
            self.root.withdraw()
        else:
            self._quit()

    def _quit(self):
        if self.tray_icon:
            self.tray_icon.stop()
        self.root.destroy()

    def _setup_systray(self):
        def img():
            i = Image.new('RGB', (64, 64), color=(47, 84, 150))
            d = ImageDraw.Draw(i)
            d.rectangle([18, 12, 46, 20], fill="white")
            d.rectangle([18, 44, 46, 52], fill="white")
            d.rectangle([26, 12, 38, 52], fill="white")
            return i

        menu = Menu(
            MenuItem("Show Dashboard", lambda i, s: self.root.after(0, self._restore_from_tray), default=True),
            Menu.SEPARATOR,
            MenuItem("Log New Project", lambda i, s: self.root.after(0, self._manual_sw_log)),
            MenuItem("Log Print/Filament", lambda i, s: self.root.after(0, self._manual_print_log)),
            MenuItem("Log Handover", lambda i, s: self.root.after(0, self._manual_handover_log)),
            Menu.SEPARATOR,
            MenuItem("Exit", lambda i, s: self.root.after(0, self._quit)),
        )
        self.tray_icon = Icon("JigWF", img(), "Jig Workflow Automation", menu)
        threading.Thread(target=self.tray_icon.run, daemon=True).start()


if __name__ == "__main__":
    try:
        App()
    except Exception as e:
        log.critical(traceback.format_exc())
        try:
            tk.Tk().withdraw()
            messagebox.showerror("Fatal Error", str(e))
        except:
            pass
